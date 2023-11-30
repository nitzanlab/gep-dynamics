#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module implementing gene expression programs (GEPs) comparator between two
timepoints, examining timepoint a as the basis for timepoint b

input:
1. timepoint A data
2. timepoint A usages matrix (H in Matan’s terminology)
3. timepoint B data

pseudocode:
0. Identify A GEPs on timepoints A and B jointly highly variable genes
1. Decompose timepoint B data de-novo and using timepoint A GEP matrix with
    0,1,2,3 additional degrees of freedom.
2. Calculate gene coefficients for all the GEPs
3. Calculate the coefficients correlation of each A GEP with all the novel GEPs
4. Calculate the usage correlation of the de-novo and pfnmf GEPs
5. Input the correlation results to the fingerprint surrogate table
6. Classify GEPs to operators according to the fingerprint

output:
1. Classification of each GEP in timepoint A according to the operator
2. Fingerprint figure for each GEP in the timepoint A decomposition.

"""
import os
import numbers
import resource
import typing

from copy import copy
from typing import Tuple, Dict, Any, List, Union

import networkx as nx
import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl import load_workbook
from sklearn.decomposition import _nmf as sknmf
from scipy.cluster import hierarchy

import gepdynamics._utils as _utils
import gepdynamics.cnmf as cnmf
from gepdynamics import pfnmf
from gepdynamics._constants import NMFEngine, Stage, NUMBER_HVG, NMF_TOLERANCE, EXCEL_FINGERPRINT_TEMPLATE


class NMFResultBase(object):
    """
    A base class for storing and manipulating results from Non-negative Matrix Factorization (NMF).

    This class should be extended by specific NMF result classes that calculate the attributes
    and implement any additional functionality specific to the type of NMF performed.

    Attributes
    ----------
    name : str
        The name identifier for the NMF result.
    loss_per_cell : np.ndarray
        Array of shape (n_cells,) containing the loss for each cell.
    rank : int
        The rank used in the NMF algorithm.
    algorithm : str
        The name of the algorithm used for NMF. Valid values are 'regular' and 'pfnmf'.
    loss : float
        The total loss computed as the sum of `loss_per_cell`.
    prog_names : list of str
        List of program names.
    norm_usages : np.ndarray
        Array of shape (n_cells, rank) with normalized usages per cell.
    prog_percentages : np.ndarray
        Array of shape (rank,) with total program usage percentages across all dataset.
    prog_labels_1l : list of str
        List of program labels in one line.
    prog_labels_2l : list of str
        List of program labels in two lines.
    gene_coefs : pd.DataFrame
        Gene coefficients for each program, shape is (n_genes, rank).
    gsea_results : list of pd.DataFrame
        Results from Gene Set Enrichment Analysis. see `gepdynamics._utils.gsea`.

    Methods
    -------
    __init__(self, name: str, loss_per_cell: np.ndarray, rank: int, algorithm: str)
        Initializes the NMFResultBase instance, extended by subclasses.

    calculate_prog_labels(self)
        Calculates the program labels and percentages. Utilized in subclasses constructors

    aggregate_results(results: List['NMFResultBase']) -> Tuple[list, list, np.ndarray, list]
        Aggregates results from a list of NMFResultBase objects for use in correlations flow chart.

    calculate_gene_coefs(self, z_scores: np.ndarray, gene_index: pd.Index)
        Calculates gene coefficients for each program based on Z-scores.

    get_top_genes(self, n: int = 100) -> pd.DataFrame
        Returns a DataFrame with the top 'n' genes based on their coefficients for each program.

    plot_loss_per_cell_histogram(self, bins: int = 25, max_value: float = 2500,
                                 saving_folder: _utils.PathLike = None,
                                 save: bool = True, show: bool = False)
        Plots and optionally saves a histogram of the loss per cell distribution.

    plot_correlations_heatmaps(self, tsc_truncation_level: int = 1000,
                               saving_folder: _utils.PathLike = None,
                               save: bool = True, show: bool = False):
        Plots heatmaps for the usage and gene coefficients correlations.

    plot_utilization_scatter(self, coordinates: np.ndarray,
                             saving_folder: _utils.PathLike = None,
                             save: bool = True, show: bool = False):
        Plots a scatter plot of utilization patterns based on given coordinates.
    """
    def __init__(self, name: str, loss_per_cell: np.ndarray, rank: int, algorithm: str):
        """
        Constructs all the necessary attributes for the NMFResultBase object.
        Most attributes are set by the subclasses constructors.

        Parameters
        ----------
        name : str
            Name identifier for the NMF result.
        loss_per_cell : np.ndarray
            Loss per cell as an array.
        rank : int
            Rank of the result.
        algorithm : str
            Name of the NMF algorithm used.
        """
        self.name = name
        self.loss_per_cell = loss_per_cell
        self.rank = rank
        self.algorithm = algorithm

        self.loss = np.sum(self.loss_per_cell)

        # Attributes added later on
        self.prog_names = None
        self.norm_usages = None
        self.prog_percentages = None
        self.prog_labels_1l = None
        self.prog_labels_2l = None
        self.gene_coefs = None
        self.gsea_results = None

    def calculate_prog_labels(self):
        """
        Calculates and assigns program total use percentages and labels to the instance.
        Program percentages are calculated based on the sum of normalized usages
        across cells. Two types of labels are created: one-line labels and two-line labels.
        """
        self.prog_percentages = self.norm_usages.sum(axis=0) * 100 / len(self.loss_per_cell)
        self.prog_labels_1l = [name + f' ({self.prog_percentages[i]:0.1f}%)' for i, name in enumerate(self.prog_names)]
        self.prog_labels_2l = [name + f'\n({self.prog_percentages[i]:0.1f}%)' for i, name in enumerate(self.prog_names)]

    @staticmethod
    def aggregate_results(results: List['NMFResultBase']) -> Tuple[list, list, np.ndarray, list]:
        """
        Aggregate results from a list of NMFResult objects, used in correaltions flow chart.

        Parameters
        ----------
        results : list
            A list of NMFResult objects.

        Returns
        -------
        Tuple[list, list, np.ndarray, list]
            A tuple containing the aggregated results:
            - joint_ranks: List of ranks from each NMFResult object.
            - joint_names: List of program names from all NMFResult objects.
            - joint_usages: Concatenated array of norm_usages from all NMFResult objects.
            - joint_labels: List of program labels from all NMFResult objects.
        """
        joint_ranks = [r.rank for r in results]
        joint_names = [name for r in results for name in r.prog_names]
        joint_labels = [label for r in results for label in r.prog_labels_1l]

        usages = [r.norm_usages for r in results]

        joint_usages = np.concatenate(usages, axis=1)

        return joint_ranks, joint_names, joint_usages, joint_labels

    def calculate_gene_coefs(self, z_scores: np.ndarray, gene_index: pd.Index):
        """
        Calculate gene coefficients for each program. Updates the `gene_coefs`
        attribute with a DataFrame of shape (n_genes, rank) for the results.

        Parameters
        ----------
        z_scores : np.ndarray
            Z-scores matrix of shape (n_genes, n_cells).
        gene_index : pd.Index
            List of gene names

        """
        self.gene_coefs = pd.DataFrame(
            _utils.fastols(self.norm_usages, z_scores).T,
            index=gene_index,
            columns=self.prog_names)

    def get_top_genes(self, n: int = 100) -> pd.DataFrame:
        """
        Return a dataframe with the top genes coefficients for each program.
        """
        return pd.DataFrame([self.gene_coefs.nlargest(n, prog).index.tolist()
                             for prog in self.prog_names], index=self.prog_names).T

    def plot_loss_per_cell_histogram(self, bins: int = 25, max_value: float = 2500,
                                     saving_folder: _utils.PathLike = None,
                                     save: bool = True, show: bool = False):
        """
        Plots a histogram of the loss per cell and optionally saves it as an image.

        Parameters:
        ----------
        bins : int, optional
            Number of bins for the histogram (default is 25).
        max_value : float, optional
            Maximum value to consider for the histogram range (default is 2500).
        saving_folder : _utils.PathLike, optional
            Path to the folder where the plot will be saved (default is None).
        save : bool, optional
            Whether to save the plot as an image file (default is True).
        show : bool, optional
            Whether to display the plot (default is False).

        Raises:
        -------
        ValueError
            If `saving_folder` is not specified when `save=True`.

        Returns:
        --------
        None
        """
        if saving_folder is None and save:
            raise ValueError('saving_folder must be specified if save=True')

        plt.hist(self.loss_per_cell, bins=bins, range=(0, min(max_value, self.loss_per_cell.max())))
        plt.title(f'{self.name} loss per cell distribution')

        if save:
            saving_folder = _utils.set_dir(saving_folder)
            plt.savefig(saving_folder.joinpath(f'{self.name}_loss_per_cell_distribution.png'))
        if show:
            plt.show()
        plt.close()

    def plot_correlations_heatmaps(self, tsc_truncation_level: int = 1000,
                                   saving_folder: _utils.PathLike = None,
                                   save: bool = True, show: bool = False):
        """
        Plots heatmaps for the usage and gene coefficients correlations.

        Usages correlation is calculated as Pearson correlation. gene coefficients
        correlation is calculated as truncated spearman correlation with a cutoff
        of `tsc_truncation_level` genes.

        Parameters:
        ----------
        tsc_truncation_level : int, optional
            Number of top ranked genes for the truncated spearman correlation
            of the gene coefficients (default is 1000).
        saving_folder : _utils.PathLike, optional
            Path to the folder where the plots will be saved (default is None).
        save : bool, optional
            Whether to save the plots as image files (default is True).
        show : bool, optional
            Whether to display the plots (default is False).

        Raises:
        -------
        ValueError
            If `saving_folder` is not specified when `save=True`.

        Returns:
        --------
        None
        """
        if saving_folder is None and save:
            raise ValueError('saving_folder must be specified if save=True')
        elif not save:
            file_genes = None
            file_usage = None
            file_H = None
        else:
            dec_folder = _utils.set_dir(saving_folder)
            file_genes = dec_folder.joinpath(f'{self.name}_correlations_gene_coefficients.png')
            file_H = dec_folder.joinpath(f'{self.name}_correlations_H.png')
            file_usage = dec_folder.joinpath(f'{self.name}_correlations_usages.png')

        # parameters for genes heatmap
        df_genes = pd.DataFrame(_utils.truncated_spearmans_correlation(
            self.gene_coefs, truncation_level=tsc_truncation_level, rowvar=False),
            index=self.prog_names, columns=self.prog_names)
        title_genes = f'{self.name} gene coefficients truncated spearman correlation'

        # parameters for H heatmap
        df_H = pd.DataFrame(np.corrcoef(self.H, rowvar=True),
                            index=self.prog_names, columns=self.prog_names)
        title_H = f'{self.name} programs matrix correlations'

        # parameters for usages heatmap
        df_usages = pd.DataFrame(np.corrcoef(self.norm_usages, rowvar=False),
                          index=self.prog_names, columns=self.prog_names)
        title_usage = f'{self.name} usages correlation'

        for df, title, file in [
            (df_genes, title_genes, file_genes),
            (df_usages, title_usage, file_usage),
            (df_H, title_H, file_H)]:
            g = sns.heatmap(df, cmap='RdYlGn', annot=True, fmt='.2f', vmin=0,
                            vmax=1, square=True, cbar_kws={"shrink": (1 - 0.01*self.rank)})
            g.set_xticklabels(g.get_xticklabels(), rotation=45, horizontalalignment='right')
            g.set_yticklabels(g.get_yticklabels(), rotation=0, horizontalalignment='right')
            g.figure.set_figwidth(5 + 0.4 * self.rank)
            g.figure.set_figheight(4 + 0.4 * self.rank)
            plt.title(title)
            plt.tight_layout()

            if save:
                plt.savefig(file)
            if show:
                plt.show()
            plt.close()

    def plot_utilization_scatter(self, coordinates: np.ndarray,
                                saving_folder: _utils.PathLike = None,
                                save: bool = True, show: bool = False):
        """
        Plots a scatter plot of utilization patterns based on given coordinates.

        Parameters:
        ----------
        coordinates : np.ndarray
            Array of shape (n_cells, 2) containing the x, y coordinates for
            plotting the gene coefficients.
        saving_folder : _utils.PathLike, optional
            Path to the folder where the plots will be saved (default is None).
        save : bool, optional
            Whether to save the plots as image files (default is True).
        show : bool, optional
            Whether to display the plots (default is False).

        Raises:
        -------
        ValueError
            If `saving_folder` is not specified when `save=True`.

        Returns:
        --------
        None
        """
        if saving_folder is None and save:
            raise ValueError('saving_folder must be specified if save=True')
        elif save:
            projections_folder = _utils.set_dir(saving_folder)

        color_map = sns.color_palette(f"light:green", as_cmap=True)

        for i in range(self.rank):
            plt.scatter(
                coordinates[:, 0], coordinates[:, 1], s=2,
                c=_utils.floats_to_colors(self.norm_usages[:, i], color_map),
                # marker=adata.uns['marker'],
                label=f'{self.prog_names[i]}')

            ax = plt.gca()

            ax.set_xlabel('UMAP 1')
            ax.set_ylabel('UMAP 2')
            ax.set_title(f'Program Utilization UMAP')

            ax.tick_params(bottom=False, left=False, labelbottom=False, labelleft=False)

            plt.tight_layout()
            plt.legend()

            if save:
                plt.savefig(projections_folder.joinpath(f'{self.prog_names[i]}.png'))
            if show:
                plt.show()
            plt.close()


class NMFResult(NMFResultBase):
    """
    A class representing the results of a standard NMF analysis.

    Inherits from NMFResultBase.

    Additional Attributes
    ----------
    W : np.ndarray
        Array of shape (n_cells, rank) containing the program usages per cell.
    H : np.ndarray
        Array of shape (n_genes, rank) containing the gene weights per program.


    Methods
    -------
    Inherits all methods from NMFResultBase.
    """
    def __init__(self, name, loss_per_cell, rank, W, H):
        """
        Initializes the NMFResult instance.

        Parameters
        ----------
        name : str
            The name of the NMF analysis.
        loss_per_cell : np.ndarray
            Array containing the loss per cell with shape (n_cells,).
        rank : int
            The rank of the factorization, corresponding to the number of components or programs.
        W : np.ndarray
            Program usage matrix obtained from NMF with shape (n_cells, rank).
        H : np.ndarray
            Gene weights matrix obtained from NMF with shape (rank, n_genes).

        Raises
        ------
        AssertionError
            If the second dimension of W does not match the first dimension of H
            and the rank, or if the first dimensions of W do not align with
            loss_per_cell.
        """
        assert W.shape[1] == H.shape[0], "The second dimension of W must match the first dimension of H"
        assert W.shape[0] == loss_per_cell.shape[0], "The first dimension of W must match the first dimension of loss_per_cell"
        assert rank == W.shape[1], "The rank must match the second dimension of W"

        super().__init__(name, loss_per_cell, rank, 'regular')
        self.W = W
        self.H = H

        self.norm_usages = W / W.sum(axis=1, keepdims=True)
        self.prog_names = [f'{name}.p{i}' for i in range(rank)]

        super().calculate_prog_labels()


class PFNMFResult(NMFResultBase):
    """
    A class representing the results of Partially Fixed Non-negative Matrix Factorization (PFNMF).

    Inherits from NMFResultBase and initializes additional attributes specific to
    the PFNMF results.

    Additional Attributes
    ----------
    W1 : np.ndarray
        Fixed program usage matrix with shape (n_cells, n_fixed_programs).
    W2 : np.ndarray
        Novel program usage matrix with shape (n_cells, n_novel_programs).
    H1 : np.ndarray
        Fixed gene weights matrix with shape (n_fixed_programs, n_genes).
    H2 : np.ndarray
        Novel gene weights matrix with shape (n_novel_programs, n_genes).


    Methods
    -------
    Inherits all methods from NMFResultBase.
    """
    def __init__(self, name, loss_per_cell, rank, W1, W2, H1, H2):
        """
        Initialize the PFNMFResult instance.

        Parameters
        ----------
        name : str
            The name of the decomposition instance.
        loss_per_cell : np.ndarray
            Array containing the loss per cell with shape (n_cells,).
        rank : int
            The total rank of the factorization, which is the sum of fixed and novel programs.
        W1 : np.ndarray
            Fixed program usage matrix with shape (n_cells, n_fixed_programs).
        W2 : np.ndarray
            Novel program usage matrix with shape (n_cells, n_novel_programs).
        H1 : np.ndarray
            Fixed gene weights matrix with shape (n_fixed_programs, n_genes).
        H2 : np.ndarray
            Novel gene weights matrix with shape (n_novel_programs, n_genes).

        Raises
        ------
        AssertionError
            If the dimensions of W1 and H1, or W2 and H2, do not align, or if the
            first dimensions of W1 do not align with W2 and loss_per_cell, or if
            the sum of program counts in W1 and W2 does not match the rank.
        """
        assert W1.shape[1] == H1.shape[0], "The second dimension of W1 must match the first dimension of H1"
        assert W2.shape[1] == H2.shape[0], "The second dimension of W2 must match the first dimension of H2"
        assert W1.shape[0] == W2.shape[0], "The first dimension of W1 must match the first dimension of W2"
        assert W1.shape[0] == loss_per_cell.shape[0], "The first dimension of W1 must match the first dimension of loss_per_cell"
        assert rank == W1.shape[1] + W2.shape[1], "The rank must match the sum of the second dimensions of W1 and W2"

        super().__init__(name, loss_per_cell, rank, 'pfnmf')
        self.W1 = W1
        self.W2 = W2
        self.H1 = H1
        self.H2 = H2
        self.H = np.concatenate([H1, H2], axis=0)

        usages = np.concatenate([W1, W2], axis=1)
        self.norm_usages = usages / usages.sum(axis=1, keepdims=True)

        self.prog_names = [f'{name}.p{i}' for i in range(W1.shape[1])]
        self.prog_names.extend([f'{name}.e{i}' for i in range(W2.shape[1])])

        super().calculate_prog_labels()


class Comparator(object):
    """
    A class for comparing two datasets (referred to as timepoints A and B) by examining the
    applicability of Gene Expression Programs (GEPs) derived from one timepoint to the decomposition
    of the other. This class is primarily used to understand how gene expressions evolve or vary
    between different biological conditions or timepoints.

    The comparison is done through various methods of Non-negative Matrix Factorization (NMF),
    including de-novo decomposition and fixed/novel program decomposition (PFNMF), enabling
    a comprehensive analysis of shared and unique gene expression patterns between the two timepoints.

    Attributes
    ----------
    adata_a : sc.AnnData
        scanpy annotated data object for timepoint A.
    adata_b : sc.AnnData
        scanpy annotated data object for timepoint B.
    usages_matrix_a : np.ndarray
        Usages matrix derived from timepoint A, used for comparative analysis.
    rank_a : int
        Rank of the NMF decomposition for timepoint A.
    results_dir : _utils.PathLike
        Directory path for saving analysis results.
    nmf_engine : NMFEngine
        Enum specifying the NMF engine to be used (e.g. sklearn and torchnmf).
    beta_loss : str
        Loss function to be used in NMF (e.g. 'kullback-leibler' and 'Frobenius').
    max_nmf_iter : int
        Maximum number of iterations for the NMF algorithm.
    max_added_rank : int
        Maximum additional rank to consider for PFNMF.
    joint_hvgs : Union[str, int]
        Highly variable genes; either a count or a key in `adata.var`.
    device : str
        Computing device for torch-based NMF engine.
    verbosity : int
        Verbosity level for logging.

    Examples
    --------
    >>> comparator = Comparator(adata_a, usages_matrix_a, adata_b, results_dir)
    >>> comparator.extract_geps_on_jointly_hvgs()
    >>> comparator.examine_adata_a_decomposition_on_jointly_hvgs(show=True)
    >>> comparator.decompose_b()

    Notes
    -----
    The Comparator class is part of a larger workflow in the analysis of gene expression data,
    particularly in the context of understanding temporal changes of expression programs.
    It should be used in conjunction with appropriate preprocessing steps and subsequent
    analysis for comprehensive insights.

    # ToDo: create log file
    # ToDo: add colors
    """
    def __init__(self,
                 adata_a: sc.AnnData,
                 usages_matrix_a: np.ndarray,
                 adata_b: sc.AnnData,
                 results_dir: _utils.PathLike,
                 nmf_engine: NMFEngine = NMFEngine.sklearn,
                 beta_loss: str = 'kullback-leibler',
                 max_nmf_iter: int = 800,
                 max_added_rank: int = 3,
                 highly_variable_genes: Union[str, int] = NUMBER_HVG,
                 tpm_target_sum: int = 100_000,
                 decomposition_normalization_method: typing.Literal['variance', 'variance_cap'] = 'variance',
                 device: str = None,
                 verbosity: int = 0
                 ):
        """
        Initialize the Comparator class.

        Parameters
        ----------
        adata_a : sc.AnnData
            Annotated data for timepoint A.
        usages_matrix_a : np.ndarray
            Usages matrix for timepoint A.
        adata_b : sc.AnnData
            Annotated data for timepoint B.
        results_dir : _utils.PathLike
            Path to the directory where results will be stored.
        nmf_engine : NMFEngine, optional
            NMF engine to use for decomposition, by default NMFEngine.sklearn.
        beta_loss : str, optional
            Beta divergence loss function, by default 'kullback-leibler'.
        max_nmf_iter : int, optional
            Maximum number of iterations for NMF, by default 800.
        max_added_rank : int, optional
            Maximum additional rank for GEPs, by default 3.
        highly_variable_genes : Union[str, int], optional
            Either the adata_a.var key for joint highly variable genes, or the
            number of automatically found jointly highly variable genes for the
             two datasets. by default `_constants.NUMBER_HVG`.
        tpm_target_sum : int, optional
            target sum of transcript per (num) normalization as part of 
            calculating the gene coefficients (prior to log transform and
            z-scoring), by default 100,000
        decomposition_normalization_method : str, optional
            Normalization method for decomposition, by default 'variance'.
        device : str, optional
            Device to use for torch-based NMF engine, by default None.
        verbosity : int, optional
            Level of verbosity for logging. Valid values are 0 (no logging),
            1 (minimal logging), and 2 (detailed logging). Defaults to 0.
        """

        self.adata_a = adata_a
        self.usages_matrix_a = usages_matrix_a
        self.rank_a = self.usages_matrix_a.shape[1]
        self.adata_b = adata_b
        self.a_sname = self.adata_a.uns['sname']
        self.b_sname = self.adata_b.uns['sname']
        self.tpm_target_sum = tpm_target_sum

        self.results_dir = _utils.set_dir(results_dir)
        self.verbosity = verbosity

        self.beta_loss = beta_loss
        self.max_nmf_iter = max_nmf_iter
        self.max_added_rank = max_added_rank
        self.nmf_engine = nmf_engine
        self.decomposition_normalization_method = decomposition_normalization_method
        self.device = device

        if isinstance(highly_variable_genes, str):
            if highly_variable_genes not in self.adata_a.var.keys():
                raise KeyError(f"adata_a.var does not contain key {highly_variable_genes}")
        elif isinstance(highly_variable_genes, int):
            assert highly_variable_genes <= self.adata_a.n_vars, \
                f"Number of highly variable genes {highly_variable_genes} " \
                f"exceeds the number of genes in adata_a ({self.adata_a.n_vars})"
        else:
            # wrong type of highly variable genes
            raise TypeError(f"highly_variable_genes must be either str or int")
        self.joint_hvgs = highly_variable_genes

        # Fields to be filled in later
        self.geps_a = None
        self.mask_b_genes = None
        self.a_result = None
        self.fnmf_result = None
        self.pfnmf_results = []
        self.denovo_results = []
        self._all_results = []

        self.fingerprints = [None] * self.rank_a

        self.stage = Stage.INITIALIZED

    def _set_torch_device(self):
        """
        Set the torch device to use for torch-based NMF engine. return the device.

        """
        if 'torch' not in globals():
            global torch
            print('trying to import torch')
            try:
                import torch
            except ImportError:
                raise ImportError("torch module is not installed. To use torchnmf engine please install torch.")

        if isinstance(self.device, torch.device):
            pass
        elif self.device is None:
            self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
            print('No torch device was specified, using', self.device)
        else:
            torch.device(self.device)
        return self.device

    def __repr__(self):
        """
        String representation of the Comparator object.
        """
        return f'Comparator(adata_a={self.a_sname}, adata_b={self.b_sname}) at' \
               f' stage {self.stage}. engine={self.nmf_engine}.'

    def save_to_file(self, filename: _utils.PathLike):
        """
        Save the Comparator object to a ".npz" file, excluding the adata fields.

        Parameters
        ----------
        filename : PathLike
            The path and filename to save the object to. Behavs like the `filename` parameter of `numpy.savez`.

        Returns
        -------
        None
            This method does not return anything.

        Notes
        -----
        This method creates a copy of the Comparator object and sets the `adata_a` and `adata_b` fields to `None`.
        The copied object is then saved to the specified file using the `.npz` format.

        Examples
        --------
        >>> comp = Comparator()
        >>> comp.save_to_file('comp_object.npz')
        """
        copy_of_self = copy(self)
        copy_of_self.adata_a = None
        copy_of_self.adata_b = None

        # save the copied object
        np.savez(filename, obj=copy_of_self)

    @staticmethod
    def load_from_file(filename: _utils.PathLike, adata_a: sc.AnnData, adata_b: sc.AnnData) -> 'Comparator':
        """
        Load a Comparator object from a file.

        Parameters
        ----------
        filename : PathLike
            The path and filename of the file to load the object from.
        adata_a : scanpy.AnnData
            The `AnnData` object associated with the `adata_a` attribute of the loaded object.
        adata_b : scanpy.AnnData
            The `AnnData` object associated with the `adata_b` attribute of the loaded object.

        Returns
        -------
        Comparator
            The loaded Comparator object.

        Notes
        -----
        This method loads a Comparator object from a file saved in the `.npz` format.
        The loaded object is assigned the provided `adata_a` and `adata_b` attributes before returning.

        Examples
        --------
        >>> comp = Comparator.load_from_file('comp_object.npz', adata_a, adata_b)
        """
        # make sure the file exists
        if not os.path.exists(filename):
            raise FileNotFoundError(f'File {filename} does not exist')

        new_instance = np.load(filename, allow_pickle=True)['obj'].item()

        assert new_instance.a_sname == adata_a.uns['sname'],\
            f"adata_a's sname ({adata_a.uns['sname']}) does not match the sname" \
            f" in the loaded object ({new_instance.a_sname})"

        assert new_instance.b_sname == adata_b.uns['sname'],\
            f"adata_b's sname ({adata_b.uns['sname']}) does not match the sname" \
            f" in the loaded object ({new_instance.b_sname})"

        new_instance.adata_a = adata_a
        new_instance.adata_b = adata_b
        return new_instance

    def _normalize_adata_for_decomposition(self, adata: sc.AnnData,
                                           method: typing.Literal['variance', 'variance_cap'] = 'variance',
                                           min_cell_per_gene_percent: float = 1.) -> np.ndarray:
        """
        Normalize adata for decomposition, fit the data type to `self.usages_matrix_a.dtype`.
        """

        normalized = _utils.subset_and_normalize_for_nmf(
            adata, subset_by=self.joint_hvgs, method=method,
            min_cells_percent=min_cell_per_gene_percent,
            dtype=self.usages_matrix_a.dtype)
        return normalized

    def _run_nmf_with_known_usages(self,
                                   usages_matrix: np.ndarray,
                                   data: np.ndarray,
                                   name: str,
                                   tens: 'torch.Tensor' = None) -> NMFResult:
        """
        Run NMF with known usages matrix
        """
        # working in the transposed notation to fit the NMF API
        #  to get the programs: data.T ~ geps.T @ usages.T

        nmf_kwargs = {'H': usages_matrix.T.copy(),
                      'update_H': False,
                      'n_components': usages_matrix.shape[1],
                      'tol': NMF_TOLERANCE,
                      'max_iter': self.max_nmf_iter,
                      'beta_loss': self.beta_loss,
                      'solver': 'mu'
                      }

        if self.nmf_engine == NMFEngine.torchnmf:
            if tens is None:
                W, H, n_iter = cnmf.nmf_torch(
                    data.T, nmf_kwargs, device=self._set_torch_device(),
                    verbose=(self.verbosity > 1))
            else:
                W, H, n_iter = cnmf.nmf_torch(
                    data.T, nmf_kwargs, tens=tens.T, verbose=(self.verbosity > 1))
        elif self.nmf_engine == NMFEngine.sklearn:
            W, H, n_iter = sknmf.non_negative_factorization(
                data.T, **nmf_kwargs, verbose=(self.verbosity > 1))
        else: # not implemented
            raise NotImplementedError(f'NMF engine {self.nmf_engine} is not implemented')

        loss_per_cell = pfnmf.calc_beta_divergence(
            data.T, W, np.zeros((W.shape[0], 0)),
            H, np.zeros((0, H.shape[1])), per_column=True)

        return NMFResult(
            name=name,
            loss_per_cell=loss_per_cell,
            rank=W.shape[1],
            W=H.T,
            H=W.T)

    def extract_geps_on_jointly_hvgs(self, min_cells_per_gene: int = 5,
                                     coefs_variance_normalization: typing.Literal[None, 'variances_norm'] = None):
        """
        Identify Gene Expression Programs (GEPs) on genes that are highly variable
        across both timepoints A and B.

        This method can operate in two modes depending on the type of `self.joint_hvgs`:
        1. If `self.joint_hvgs` is an integer, it selects that number of top highly variable genes based on
           the combined variability ranking from both timepoints.
        2. If `self.joint_hvgs` is a string, it is assumed to be a key in `adata_a.var`,
         and the method uses the corresponding variable genes already annotated in `adata_a`.

        After selecting the highly variable genes, the method performs normalization
        and re-decomposes timepoint A using the known usages matrix, this time
        on the jointly highly variable genes.

        Parameters
        ----------
        min_cells_per_gene : int, optional
            Minimum number of cells in which a gene must be expressed to be considered in the analysis.
            This parameter is only used if `self.joint_hvgs` is an integer. Defaults to 5.
        coefs_variance_normalization : typing.Literal[None, 'variances_norm'], optional
            Whether to normalize the gene coefficients by the normalized variances
            of the genes, by default None.

        Raises
        ------
        RuntimeError
            If the method is called when the object is not in the INITIALIZED stage

        Notes
        -----
        The method sets the object to the PREPARED stage upon successful completion.
        """
        # assert state is INITIALIZED
        if self.stage != Stage.INITIALIZED:
            raise RuntimeError(f"Stage is {self.stage}, expected {Stage.INITIALIZED}")

        # if joint_hvgs is numeric, use it as the number of joint genes
        if isinstance(self.joint_hvgs, int):
            genes_subset = (self.adata_a.var.n_cells >= min_cells_per_gene) & (
                self.adata_b.var.n_cells >= min_cells_per_gene)
            hvg_a = sc.pp.highly_variable_genes(
                self.adata_a[:, genes_subset], flavor='seurat_v3', n_top_genes=100_000, inplace=False)
            hvg_b = sc.pp.highly_variable_genes(
                self.adata_b[:, genes_subset], flavor='seurat_v3', n_top_genes=100_000, inplace=False)
            joint_hvg_rank = hvg_a.highly_variable_rank + hvg_b.highly_variable_rank
            self.joint_hvgs = joint_hvg_rank.sort_values()[: self.joint_hvgs].index
        elif isinstance(self.joint_hvgs, str):
            self.joint_hvgs = self.adata_a.var.index[self.adata_a.var[self.joint_hvgs]]
        elif isinstance(self.joint_hvgs, pd.core.indexes.base.Index):
            pass
        else:
            raise TypeError(f"joint_hvgs must be either int or a key in `adata_a.var`")

        if self.verbosity > 0:
            print(f'Extracting A GEPs on jointly highly variable genes')

        X_a = self._normalize_adata_for_decomposition(
            self.adata_a, self.decomposition_normalization_method, -1)

        self.a_result = self._run_nmf_with_known_usages(
            self.usages_matrix_a, X_a, 'A')

        self.geps_a = self.a_result.H / np.linalg.norm(self.a_result.H, ord=2, axis=1, keepdims=True)

        # create mask of low expressed genes
        min_cells = (self.adata_b.shape[0] * 1. / 100)
        X = self.adata_b[:, self.joint_hvgs].X
        if isinstance(X, _utils.sparse.spmatrix):
            X = X.toarray()
        self.mask_b_genes = np.count_nonzero(X, axis=0) > min_cells

        # calculate gene coefs for each GEP
        if coefs_variance_normalization is None:
            self._calculate_gene_coefficients(
                self.adata_a, [self.a_result], self.tpm_target_sum)
        elif coefs_variance_normalization == 'variances_norm':
            var_norm = sc.pp.highly_variable_genes(
                self.adata_a, flavor='seurat_v3', n_top_genes=5000,
                inplace=False)['variances_norm'].values
            self._calculate_gene_coefficients(
                self.adata_a, [self.a_result], self.tpm_target_sum,
                target_variance=var_norm)

        self.stage = Stage.PREPARED

    def examine_adata_a_decomposition_on_jointly_hvgs(self, hist_bins: int = 25,
                                                      hist_max_value: float = 2500,
                                                      show: bool = False, save: bool = True):
        """
        Examine the decomposition results of timepoint A by plotting various visualizations.

        This method generates and optionally displays/saves a heatmap of correlations,
        a clustermap of normalized usages, and a histogram of the loss per cell.

        Parameters
        ----------
        hist_bins : int, optional
            Number of bins for the loss per cell histogram. Defaults to 25.
        hist_max_value : float, optional
            Maximum value for the histogram range. Defaults to 2500.
        show : bool, optional
            Whether to display the plots. Defaults to False.
        save : bool, optional
            Whether to save the plots as image files. Defaults to True.
            plots are saved under `self.results_dir`/decompositions.

        Raises
        ------
        RuntimeError
            If the method is called before extracting GEPs on jointly highly variable genes.
        """
        if self.stage == Stage.INITIALIZED:
            raise RuntimeError('Must extract GEPs on jointly HVGs first')

        res = self.a_result
        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        # correlations heatmap
        res.plot_correlations_heatmaps(saving_folder=dec_folder, show=show, save=save)

        # clustermap of normalized usages
        title = f"{self.a_sname} normalized usages of " \
                f"original GEPs, k={res.rank}"

        row_colors = _utils.expand_adata_row_colors(self.adata_a, pd.Series(
            _utils.floats_to_colors(res.loss_per_cell, cmap='RdYlGn_r', vmax=(
                2 * np.median(res.loss_per_cell))),
            name='residual', index=self.adata_a.obs.index))

        col_labels = res.prog_labels_2l if res.rank < 20 else res.prog_labels_1l

        un_sns = _utils.plot_usages_norm_clustermaps(
            self.adata_a, normalized_usages=res.norm_usages,
            columns=col_labels, title=title, show=show,
            sns_clustermap_params={'row_colors': row_colors})

        if save:
            un_sns.savefig(dec_folder.joinpath(f'{res.name}_normalized_usages_clustermap.png'))
        plt.close()

        # loss per cell histogram
        plt.hist(res.loss_per_cell, bins=hist_bins, range=(0, min(hist_max_value, res.loss_per_cell.max())))
        plt.title(f'{res.name} loss per cell distribution')

        if save:
            plt.savefig(dec_folder.joinpath(f'{res.name}_loss_per_cell_distribution.png'))
        if show:
            plt.show()
        plt.close()

    def decompose_b(self, repeats: int = 1,
                    precalculated_denovo_usage_matrices: List[np.ndarray] = None,
                    coefs_variance_normalization: typing.Literal[None, 'variances_norm'] = None):
        """
        Decompose timepoint B data de-novo and using timepoint A GEP matrix with additional degrees of freedom.

        This method handles both de-novo decomposition and decomposition using timepoint A GEPs
        (with varying degrees of freedom). It updates the object state with the results of decomposition.

        Parameters
        ----------
        repeats : int, optional
            Number of times the decomposition should be repeated. Defaults to 1.
            # TBD: add support for repeats > 1 in de-novo and fnmf (requires random initialization)
        precalculated_denovo_usage_matrices : List[np.ndarray], optional
            Pre-calculated usage matrices for de-novo decomposition, if available. Defaults to None.
        coefs_variance_normalization : typing.Literal[None, 'variances_norm'], optional
            Whether to normalize the gene coefficients by the normalized variances
            of the genes, by default None.

        Raises
        ------
        RuntimeError
            If the method is called before extracting GEPs on jointly highly variable genes.
        NotImplementedError
            If an NMF engine other than 'sklearn' or 'torchnmf' is used.

        Notes
        -----
        Updates the object state to DECOMPOSED upon successful completion.

        """
        if self.stage == Stage.INITIALIZED:
            raise RuntimeError('Must extract GEPs on jointly HVGs first')

        X_b = self._normalize_adata_for_decomposition(
            self.adata_b, self.decomposition_normalization_method, -1)

        if self.nmf_engine == NMFEngine.sklearn:
            if self.verbosity > 0:
                print('Decomposing B using A GEPs and no additional GEPs')
            self._decompose_b_fnmf(X_b)

            if self.verbosity > 0:
                print('Decomposing B de-novo')
            self._decompose_b_denovo(X_b, usages_matrices=precalculated_denovo_usage_matrices)

        elif self.nmf_engine == NMFEngine.torchnmf:
            device = self._set_torch_device()
            tens = torch.tensor(X_b).to(device)

            if self.verbosity > 0:
                print('Decomposing B using A GEPs and no additional GEPs')
            self._decompose_b_fnmf(X_b, tens)

            if self.verbosity > 0:
                print('Decomposing B de-novo')
            self._decompose_b_denovo(X_b, tens, usages_matrices=precalculated_denovo_usage_matrices)

            del tens

        else:
            # not implemented
            raise NotImplementedError(f'nmf engine {self.nmf_engine} not implemented')

        if self.verbosity > 0:
            print(f'Decomposing B using A GEPs and up to {self.max_added_rank} additional GEPs')
        self._decompose_b_pfnmf(X_b, repeats)

        self._all_results = self.denovo_results + [self.fnmf_result] + self.pfnmf_results

        # calculate gene coefs for each GEP
        if coefs_variance_normalization is None:
            self._calculate_gene_coefficients(
                self.adata_b, self._all_results, self.tpm_target_sum)
        elif coefs_variance_normalization == 'variances_norm':
            var_norm = sc.pp.highly_variable_genes(
                self.adata_b, flavor='seurat_v3', n_top_genes=5000,
                inplace=False)['variances_norm'].values
            self._calculate_gene_coefficients(
                self.adata_b, self._all_results, self.tpm_target_sum,
                target_variance=var_norm)

        self.stage = Stage.DECOMPOSED

    def _decompose_b_fnmf(self, X_b: np.ndarray, tens: 'torch.Tensor' = None):
        """
        Decompose timepoint B data using timepoint A GEPs and no additional GEPs

        Parameters
        ----------
        """

        # X_b ~ W @ geps_a
        nmf_kwargs = {
            'H': self.geps_a[:, self.mask_b_genes].copy(),
            'update_H': False,
            'n_components': self.rank_a,
            'tol': NMF_TOLERANCE,
            'max_iter': self.max_nmf_iter,
            'beta_loss': self.beta_loss,
            'solver': 'mu'}

        self.fnmf_result = self._run_nmf(
            X_b, nmf_kwargs, self.a_sname, tens, verbose=self.verbosity)

    def _decompose_b_denovo(self, X_b: np.ndarray, tens: 'torch.Tensor' = None,
                            usages_matrices: List[np.ndarray] = None):
        """
        Decompose timepoint B data de-novo

        # TBD: add support for repeats > 1 in de-novo
        Parameters
        ----------
        """
        self.denovo_results = []

        if usages_matrices is None:
            for added_rank in range(self.max_added_rank + 1):
                rank = self.rank_a + added_rank

                if self.verbosity > 0:
                    print(f'Decomposing B de-novo, rank={rank}')

                nmf_kwargs={
                    'n_components': rank,
                    'tol': NMF_TOLERANCE,
                    'max_iter': self.max_nmf_iter,
                    'beta_loss': self.beta_loss,
                    'solver': 'mu'
                   }

                self.denovo_results.append(self._run_nmf(
                    X_b, nmf_kwargs, f'dn_{rank}', tens, verbose=self.verbosity))
        else:
            assert len(usages_matrices) == self.max_added_rank + 1, \
                f'Expected {self.max_added_rank + 1} usage matrices, got {len(usages_matrices)}'
            for added_rank in range(self.max_added_rank + 1):
                rank = self.rank_a + added_rank

                assert usages_matrices[added_rank].shape[1] == rank, \
                    f'Expected usage matrix rank {rank}, got {usages_matrices[added_rank].shape[1]}'

                if self.verbosity > 0:
                    print(f'Decomposing B de-novo with known usages, rank={rank}')

                self.denovo_results.append(self._run_nmf_with_known_usages(
                    usages_matrices[added_rank], X_b, f'dn_{rank}', tens)
                )

    def _decompose_b_pfnmf(self, X_b: np.ndarray, repeats: int = 1):
        """
        Decompose timepoint B data using timepoint A GEP matrix with 1,2,3 additional degrees of freedom.

        Parameters
        ----------

        """
        # pfnmf is written for constant W_1, so we will transpose as needed:
        # X_b ~ W_1 @ geps_a + W_2 @ H_2  <--> X_b.T ~ geps_a.T @ W_1.T + H_2.T @ W_2.T

        self.pfnmf_results = []

        for added_rank in range(1, self.max_added_rank + 1):
            if self.verbosity:
                print(f"Working on added rank = {added_rank}")

            name = self.a_sname + f'e{added_rank}'
            rank = self.rank_a + added_rank

            best_loss = np.infty

            for repeat in range(repeats):
                w1, h1, w2, h2, n_iter = pfnmf.pfnmf(
                    X_b.T, self.geps_a[:, self.mask_b_genes].T, rank_2=added_rank,
                    beta_loss=self.beta_loss, tol=NMF_TOLERANCE,
                    max_iter=self.max_nmf_iter, verbose=(self.verbosity > 1))

                loss_per_cell = pfnmf.calc_beta_divergence(
                    X_b.T, w1, w2, h1, h2, self.beta_loss, per_column=True)

                final_loss = np.sum(loss_per_cell)

                if final_loss <= best_loss:
                    best_loss = final_loss

                    # transposing back to original orientation
                    best_result = PFNMFResult(
                        name, loss_per_cell, rank, h1.T, h2.T, w1.T, w2.T)

                    if self.verbosity:
                        print(f"repeat {repeat}, after {n_iter} iterations reached "
                              f"error = {final_loss: .1f}")

            self.pfnmf_results.append(best_result)

    def print_errors(self):
        """
        Print the errors of the decompositions

        """
        if self.stage in (Stage.INITIALIZED, Stage.PREPARED):
            raise RuntimeError('Must decompose before printing errors')

        for result in self._all_results:
            print(f"{result.name} error = {result.loss: .1f}")

    def examine_adata_b_decompositions(self, max_kde_value: float = 2000,
                                       hist_bins: int = 25, hist_max_value: float = 2500,
                                       show: bool = False, save: bool = True):
        """
        Plot the results of the decompositions of timepoint B data

        plots:
        1. Usages and gene coefficients correlations for each decomposition
        2. Histogram of the loss per cell for each decomposition separatly
        3. Violin plots of the loss per cell for the decompositions,
        4. Joint Kernel Density Estimation of the loss per cell.

        Parameters
        ----------
        hist_bins : int, optional
            Number of bins for the histogram. Defaults to 20.
        hist_max_value: float, optional
            maximum value of x in the histograms. Defaults to 2500.
        max_kde_value: float, optional
            maximum value of x in the KDE comparison. Defaults to 2000.
        show : bool, optional
            Whether to display the plot. Defaults to False.
        save : bool, optional
            Whether to save the plot as an image. Defaults to True.
        """
        if self.stage in (Stage.INITIALIZED, Stage.PREPARED):
            raise RuntimeError('Must decompose before examining decompositions')

        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        for result in self._all_results:
            result.plot_correlations_heatmaps(
                saving_folder=dec_folder, show=show, save=save)
            result.plot_loss_per_cell_histogram(
                bins=hist_bins, max_value=hist_max_value,
                saving_folder=dec_folder, show=show, save=save)

        self._plot_loss_per_cell_histograms(max_kde_value, show, save)
        self._plot_usages_clustermaps(show, save)
        self._plot_utilization_scatters(show, save)

        self.plot_correlations_flow_chart([self.fnmf_result, *self.pfnmf_results, *self.denovo_results[::-1]],
                                          usage_corr_threshold = 0.25,
                                          tsc_threshold = 0.25,
                                          show=show, save=save)

    def _plot_loss_per_cell_histograms(self, max_kde_value: float = 2000, show: bool = False, save: bool = True):
        """
        Plot the histogram of the loss per cell for each decomposition,
        violin plots of the loss per cell for the decomposition together,
        and the KDE of the loss per cell for the decomposition together.

        Parameters
        ----------
        max_kde_value: float, optional
            maximum value of x in the KDE comparison. Defaults to 2000.
        show : bool, optional
            Whether to display the plot. Defaults to False.
        save : bool, optional
            Whether to save the plot as an image. Defaults to True.
        """
        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        df = pd.concat([pd.Series(res.loss_per_cell, name=res.name) for res in self._all_results], axis=1)
        df = pd.melt(df, var_name='decomposition', value_name='loss per cell')
        sns.violinplot(x='decomposition', y='loss per cell', data=df)

        plt.title('loss per cell')
        plt.setp(plt.gca().get_xticklabels(), rotation=45)
        if save:
            plt.savefig(dec_folder.joinpath('loss_per_cell_violin.png'))
        if show:
            plt.show()
        plt.close()

        _utils.sns.kdeplot(df, hue="decomposition", x="loss per cell", clip=(0, max_kde_value))
        plt.tight_layout()
        if save:
            plt.savefig(dec_folder.joinpath('loss_per_cell_kde.png'))
        if show:
            plt.show()
        plt.close()

    def _get_title(self, res: NMFResult):
        """
        helper function to get a title for clustermap/violin plot

        """
        if 'dn' in res.name:
            return f"{self.b_sname} %s of " \
                    f"de-novo GEPs, k={res.rank}"
        elif res.rank == self.rank_a:
            return f"{self.b_sname} %s of " \
                    f"{self.a_sname} GEPs (rank={res.rank})"
        elif res.algorithm == 'pfnmf':
            return f"{self.b_sname} %s of " \
                    f"{self.a_sname} GEPs + " \
                    f"{res.rank - self.rank_a} novel GEPs"

    def _plot_usages_clustermaps(self, show: bool = False, save: bool = True):
        """
        Plot the usages of the decomposed GEPs as clustermaps
        """
        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        for res in self._all_results:
            title = self._get_title(res) % 'normalized usages'

            row_colors = _utils.expand_adata_row_colors(self.adata_b, pd.Series(
                _utils.floats_to_colors(res.loss_per_cell, cmap='RdYlGn_r', vmax=(
                2 * np.median(res.loss_per_cell))),
                name='residual', index=self.adata_b.obs.index))

            col_labels = res.prog_labels_2l if res.rank < 20 else res.prog_labels_1l

            un_sns = _utils.plot_usages_norm_clustermaps(
                self.adata_b, normalized_usages=res.norm_usages,
                columns=col_labels, title=title, show=show,
                sns_clustermap_params={'row_colors': row_colors})

            if save:
                un_sns.savefig(dec_folder.joinpath(f'{res.name}_normalized_usages_clustermap.png'))

            plt.close()

    # plot usages violin plots for all the decompositions
    def plot_usages_violin(self, group_by_key: str, group_by_string: str = None,
                           show: bool = False, save: bool = True):
        """
        Plot the usages of the decomposed GEPs as violin plots

        Parameters
        ----------
        group_by_key : str
            The key in `adata.obs` to group the cells by.
        group_by_string : str, optional
            The string describing the grouping in the plot. Defaults to `group_by_key`.
        show : bool, optional
            Whether to display the plot. Defaults to False.
        save : bool, optional
            Whether to save the plot as an image. Defaults to True.
        """
        if group_by_string is None:
            group_by_string = group_by_key

        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        res = self.a_result
        title = (f"{self.a_sname} {group_by_string} violin plot of original GEPs"
                 f", k={res.rank}")
        _utils.plot_usages_norm_violin(
            self.adata_a, group_by_key=group_by_key, norm_usages=res.norm_usages,
            prog_names=res.prog_names, title=title, show=show,
            save_path=dec_folder.joinpath(
                f'{res.name}_normalized_usages_violin_plots.png'))

        for res in self._all_results:
            title = self._get_title(res) % f'{group_by_string} violin plot'
            _utils.plot_usages_norm_violin(
                self.adata_b, group_by_key=group_by_key,
                norm_usages=res.norm_usages,
                prog_names=res.prog_names,
                title=title,
                save_path=dec_folder.joinpath(
                    f'{res.name}_normalized_usages_violin_plots.png'),
                show=show)

    # plot utilization scatters for all the decompositions
    def _plot_utilization_scatters(self, show: bool = False, save: bool = True):
        """
        Plot the utilization scatters for all the decompositions
        """

        projections_folder = _utils.set_dir(self.results_dir.joinpath('programs_projections'))

        self.a_result.plot_utilization_scatter(
            self.adata_a.obsm['X_umap'],
            saving_folder=projections_folder, show=show, save=save)

        for res in self._all_results:
            res.plot_utilization_scatter(
                self.adata_b.obsm['X_umap'],
                saving_folder=projections_folder, show=show, save=save)


    def plot_decomposition_comparisons(self, show: bool = False,
                                       save: bool = True, max_cell_loss: float = 2000):
        """
        Plot comparisons of pfnmf and de-novo decompositions with similar ranks

        """
        if self.stage in (Stage.INITIALIZED, Stage.PREPARED):
            raise RuntimeError('Must decompose comparing decompositions')

        comp_folder = _utils.set_dir(self.results_dir.joinpath('comparisons'))

        for dn_res, pf_res in zip(self.denovo_results, [self.fnmf_result, *self.pfnmf_results]):
            # usages clustermap
            title = f"{self.b_sname} normalized usages of de-novo " \
                    f"GEPs and {self.a_sname} GEPs + {pf_res.rank - self.rank_a} novel GEPs"

            joint_usages = np.concatenate([dn_res.norm_usages, pf_res.norm_usages], axis=1)
            if dn_res.rank < 10:
                joint_labels = dn_res.prog_labels_2l + pf_res.prog_labels_2l
            else:
                joint_labels = dn_res.prog_labels_1l + pf_res.prog_labels_1l
            joint_colors = ['#2ca02c'] * dn_res.rank + ['#d62728'] * pf_res.rank

            un_sns = _utils.plot_usages_norm_clustermaps(
                self.adata_b, normalized_usages=joint_usages, columns=joint_labels,
                title=title, show=show, sns_clustermap_params={'col_colors': joint_colors})

            if save:
                un_sns.savefig(comp_folder.joinpath(f'{dn_res.name}_vs_{pf_res.name}_normalized_usages_clustermap.png'))

            plt.close()

            # loss per cell scatter
            plt.plot([0, max_cell_loss], [0, max_cell_loss], 'k-')
            plt.scatter(dn_res.loss_per_cell, pf_res.loss_per_cell, s=2,
                        c=_utils.get_single_row_color_from_adata(self.adata_b))

            plt.xlabel(f'{dn_res.name} loss per cell')
            plt.xlim((0, max_cell_loss*2))

            plt.ylabel(f'{pf_res.name} loss per cell')
            plt.ylim((0, max_cell_loss*2))

            plt.title('Comparison of loss per cell using two decompositions')

            if save:
                plt.savefig(comp_folder.joinpath(f'{dn_res.name}_vs_{pf_res.name}_loss_per_cell.png'))
            if show:
                plt.show()
            plt.close()

    def plot_correlations_flow_chart(self, results: List['NMFResultBase'],
                                     usage_corr_threshold: float = 0.3,
                                     tsc_threshold: float = 0.3,
                                     tsc_truncation_level: int = 1000,
                                     show: bool = False, save: bool = True):
        """
        Plot a flow chart of the correlations between the decompositions

        """
        if self.stage in (Stage.INITIALIZED, Stage.PREPARED):
            raise RuntimeError('Must decompose before plotting correlations')

        comparisons_folder = _utils.set_dir(self.results_dir.joinpath('comparisons'))

        names_list = [res.name for res in results]
        ks, joint_names, joint_usages, joint_labels = NMFResultBase.aggregate_results(results)

        prog_names_dict = {res.name: res.prog_names for res in results}

        # usages flow graph
        usages_title = f'Correlations flow chart of usages'
        usages_filename = f'flow_chart_usages_{"_".join(names_list)}.png'
        usages_adjacency = self._get_ordered_adjacency_matrix(
            np.corrcoef(joint_usages.T), joint_names, ks, usage_corr_threshold, verbose = self.verbosity)

        # genes flow graph
        genes_title = f'Correlations flow chart of gene coefficients'
        genes_filename = f'flow_chart_genes_{"_".join(names_list)}.png'

        tsc = _utils.truncated_spearmans_correlation(pd.concat(
            [res.gene_coefs for res in results], axis = 1),
            truncation_level = tsc_truncation_level, rowvar = False)

        genes_adjacency = self._get_ordered_adjacency_matrix(
            tsc, joint_names, ks, tsc_threshold, verbose = self.verbosity)

        for adjacency, title, filename in [(usages_adjacency, usages_title, usages_filename),
                                           (genes_adjacency, genes_title, genes_filename)]:
            fig = self._plot_layered_correlation_flow_chart(
                names_list, adjacency, prog_names_dict, title)

            if save:
                fig.savefig(comparisons_folder.joinpath(filename))
            if show:
                plt.show()
            plt.close()

    def plot_marker_genes_heatmaps(self, marker_genes: List[str],
                                   show: bool = False, save: bool = True):
        """
        Plot heatmaps of the marker genes coefficients for each decomposition
        """
        dec_folder = _utils.set_dir(self.results_dir.joinpath('decompositions'))

        for res in [self.a_result, *self._all_results]:
            sns.heatmap(res.gene_coefs.loc[marker_genes].T,
                        cmap='coolwarm', vmin=-2, vmax=2)

            plt.title(f'Marker genes coefficients for {res.name}')
            plt.tight_layout()

            if save:
                plt.savefig(dec_folder.joinpath(f'{res.name}_marker_genes_heatmap.png'))
            if show:
                plt.show()
            plt.close()

    def run_gsea(self, n_top_genes: int = 1000, gprofiler_kwargs: Dict[str, Any] = None,
                 gene_ids_column_number: int = 0):
        """
        Run gene set enrichment analysis on all the decomposed GEPs

        Parameters:
        ----------
        n_top_genes : int, optional
            Number of top-ranked genes to consider for ranked GSEA (default is 1000).
        gprofiler_kwargs : Dict[str, Any], optional
            Additional keyword arguments to be passed to `MyGProfiler` class (default is None).
        gene_ids_column_number : int, optional
            Column number in `adata_a.var` representing gene IDs (default is 0).

        Returns:
        --------
        None
        """
        if self.stage in (Stage.INITIALIZED, Stage.PREPARED):
            raise RuntimeError('Must decompose before running GSEA on decompositions')

        gene_IDs_column_name = self.adata_a.var.columns[gene_ids_column_number]

        programs_top_genes_dir = _utils.set_dir(self.results_dir.joinpath('programs_top_genes'))

        for res in [self.a_result, *self._all_results]:
            top_genes_df = res.get_top_genes(n_top_genes)
            top_genes_df.to_csv(programs_top_genes_dir.joinpath(f'{res.name}_top_genes.csv'))

        if gprofiler_kwargs is None:
            gprofiler_kwargs = {}
        gp = _utils.MyGProfiler(**gprofiler_kwargs)

        # GSEA on A decomposition GEPs
        var_names_gt_2 = sc.pp.filter_genes(self.adata_a.X, min_cells=2)[0]
        background_gene_ids = self.adata_a.var.loc[var_names_gt_2, gene_IDs_column_name].to_list()

        self._run_gsea_on_nmf_result(
            self.a_result, background_gene_ids, gp, self.adata_a.var,
            n_top_genes, gene_IDs_column_name)

        # GSEA onB decompositions GEPs
        var_names_gt_2 = sc.pp.filter_genes(self.adata_b.X, min_cells=2)[0]
        background_gene_ids = self.adata_b.var.loc[var_names_gt_2, gene_IDs_column_name].to_list()

        for res in self._all_results:
            self._run_gsea_on_nmf_result(
                res, background_gene_ids, gp, self.adata_b.var,
                n_top_genes, gene_IDs_column_name)

    def _run_gsea_on_nmf_result(
            self, res: NMFResultBase, background_gene_ids: List[str],
            gp: _utils.MyGProfiler, adata_var: pd.DataFrame, n_top_genes: int = 1000,
            gene_ids_column_name: str = 'name'):
        """
        Run gene set enrichment analysis on a specific NMF result.

        Parameters:
        ----------
        res : NMFResultBase
            NMF result on which to perform GSEA.
        background_gene_ids : List[str]
            List of background gene IDs for enrichment analysis.
        gp : _utils.MyGProfiler
            Instance of `MyGProfiler` class for GSEA.
        adata_var : pd.DataFrame
            AnnData object var attribute.
        n_top_genes : int, optional
            Number of top-ranked genes to consider for ranked GSEA (default is 1000).
        gene_ids_column_name : str, optional
            Column name in `adata_var` representing gene IDs (default is 'name').

        Returns:
        --------
        None
        """
        if self.verbosity > 1:
            print(f'Running gene set enrichment analysis on {res.name} decomposition GEPs')

        gsea_dir = _utils.set_dir(self.results_dir.joinpath('programs_gsea'))

        res.gsea_results = [None] * res.rank

        for index in range(res.rank):
            ordered_var_names = res.gene_coefs.nlargest(
                columns=[res.prog_names[index]], n=n_top_genes).index
            ordered_genesSymbol = adata_var.loc[ordered_var_names, gene_ids_column_name].to_list()

            res.gsea_results[index] = gp.profile(
                ordered_genesSymbol, ordered=True, background=background_gene_ids)

            res.gsea_results[index].to_csv(
                gsea_dir.joinpath(f"{res.prog_names[index]}.csv"))

    def calculate_fingerprints(self, truncation_level: int = 1000):
        """
        # ToDo: plot fingerprints as heatmaps (currently using csv to excel)

        """
        corr_a_vs_denovo = self._calculate_gene_coefs_tsc_a_denovo(truncation_level)
        corr_denovo_vs_pfnmf = self._calculate_corr_denovo_pfnmf(truncation_level)
        corr_a_vs_pfnmf = self._calculate_gene_coefs_tsc_a_pfnmf(truncation_level)

        for i in range(self.rank_a):
            df = corr_a_vs_denovo[i: i + 1].copy().reset_index(drop=True)
            for j in range(self.max_added_rank + 1):
                df.loc[j + 1] = corr_denovo_vs_pfnmf[j].loc[i]
            df.insert(1, 'A_coef_tsc', [np.nan, *corr_a_vs_pfnmf[:, i]])
            # df.loc[-1] = corr_a_vs_denovo.loc[i]

            self.fingerprints[i] = df.reset_index(drop=True)

        # save to file:
        csv_path = self.results_dir.joinpath(f"fingerprints_{self.a_sname}_{self.b_sname}.csv")
        excel_template = load_workbook(EXCEL_FINGERPRINT_TEMPLATE)
        xlsx_path = self.results_dir.joinpath(f"fingerprints_{self.a_sname}_{self.b_sname}.xlsx")

        separator = pd.DataFrame({'Separator': ['']})

        # saving to csv
        self.fingerprints[0].to_csv(csv_path, index=False)
        for i in range(1, self.rank_a):
            # separating line between the tables
            separator.to_csv(csv_path, mode='a', index=False, header=False)
            self.fingerprints[i].to_csv(csv_path, mode='a', index=False)

        # saving to excel
        line_offset = 0
        for index, df in enumerate(self.fingerprints):
            for col in range(0, df.shape[1]):
                excel_template.active.cell(line_offset + 1, col + 1).value = df.columns[col]

                for row in range(0, df.shape[0]):
                    excel_template.active.cell(line_offset + row + 2, col + 1).value = df.iloc[row, col]
            line_offset += df.shape[0] + 2
        excel_template.save(xlsx_path)
        excel_template.close()

    def _calculate_gene_coefs_tsc_a_denovo(self, truncation_level: int = 1000):
        """
        Calculate the TSC between the gene coefs of timepoint A and de-novo coefs
        """
        dn_res = self.denovo_results[0]

        df = pd.DataFrame(columns=['pfnmf_gep', 'dn_coef_gep', 'dn_coef_tsc',
                                   'dn_usage_gep', 'dn_usage_corr',
                                   'dn_coef_gep_2nd', 'dn_coef_tsc_2nd']
                          ).astype({'dn_coef_tsc': float, 'dn_usage_corr': float})

        df.pfnmf_gep = self.a_result.prog_labels_1l

        tsc = _utils.truncated_spearmans_correlation(pd.concat(
            [self.a_result.gene_coefs, dn_res.gene_coefs], axis=1),
            truncation_level=truncation_level, rowvar=False)[:self.rank_a, self.rank_a:]

        argmax_geps = np.argsort(tsc, axis=1)[:, -1]
        df.dn_coef_gep = [dn_res.prog_labels_1l[i] for i in argmax_geps]
        df.dn_coef_tsc = tsc[np.arange(self.rank_a), argmax_geps]

        arg_second_max_geps = np.argsort(tsc, axis=1)[:, -2]
        df.dn_coef_gep_2nd = [dn_res.prog_labels_1l[i] for i in arg_second_max_geps]
        df.dn_coef_tsc_2nd = tsc[np.arange(self.rank_a), arg_second_max_geps]

        return df

    def _calculate_gene_coefs_tsc_a_pfnmf(self, truncation_level: int = 1000):
        """
        Calculate the TSC between the gene coefs of timepoint A and f/pfNMF coefs
        """
        output = np.zeros((self.max_added_rank + 1, self.rank_a))
        for i in range(self.rank_a):
            data = pd.concat([res.gene_coefs[res.prog_names[i]] for res in [
                self.a_result, self.fnmf_result, *self.pfnmf_results]], axis=1)
            output[:, i] = _utils.truncated_spearmans_correlation(
                data, truncation_level, rowvar=False)[1:, 0]

        return output

    def _calculate_corr_denovo_pfnmf(self, truncation_level: int = 1000):
        """
        Calculate the comparisons between de-novo and f/pfNMF GEPs.
        Performs GEP usages correlation  and TSC on the gene coefficients

        for each rank of decomposition outputs a dataframe with the columns:
        pfnmf_gep, dn_coef_gep, dn_coef_tsc, dn_usage_gep, dn_usage_corr
        """
        output = []
        for dn_res, pf_res in zip(self.denovo_results, [self.fnmf_result, *self.pfnmf_results]):
            df = pd.DataFrame(columns=['pfnmf_gep', 'dn_coef_gep', 'dn_coef_tsc',
                                       'dn_usage_gep', 'dn_usage_corr',
                                       'dn_coef_gep_2nd', 'dn_coef_tsc_2nd']
                              ).astype({'dn_coef_tsc': float, 'dn_usage_corr': float})
            df.pfnmf_gep = pf_res.prog_labels_1l[:self.rank_a]

            # gene coefficients TSC:
            tsc = _utils.truncated_spearmans_correlation(pd.concat(
                [pf_res.gene_coefs.iloc[:, :self.rank_a], dn_res.gene_coefs], axis=1),
                truncation_level=truncation_level, rowvar=False)[:self.rank_a, self.rank_a:]

            argmax_geps = np.argmax(tsc, axis=1)
            df.dn_coef_gep = [dn_res.prog_labels_1l[i] for i in argmax_geps]
            df.dn_coef_tsc = tsc[np.arange(self.rank_a), argmax_geps]

            arg_second_max_geps = np.argsort(tsc, axis=1)[:, -2]
            df.dn_coef_gep_2nd = [dn_res.prog_labels_1l[i] for i in arg_second_max_geps]
            df.dn_coef_tsc_2nd = tsc[np.arange(self.rank_a), arg_second_max_geps]

            # usage correlation:
            usage_corr = np.corrcoef(np.concatenate(
                [pf_res.norm_usages[:, :self.rank_a], dn_res.norm_usages], axis=1),
                rowvar=False)[:self.rank_a, self.rank_a:]

            argmax_geps = np.argmax(usage_corr, axis=1)
            df.dn_usage_gep = [dn_res.prog_labels_1l[i] for i in argmax_geps]
            df.dn_usage_corr = usage_corr[np.arange(self.rank_a), argmax_geps]

            output.append(df)
        return output

    @staticmethod
    def _calculate_gene_coefficients(adata: sc.AnnData,
                                     results_list: List[NMFResultBase],
                                     target_sum=100_000,
                                     target_variance: typing.Union[float, np.ndarray, pd.Series] = 1.):
        """
        Calculate gene coefficients for all the newly decomposed GEPs
        """
        max_value = 10

        z_scores = sc.pp.normalize_total(adata, target_sum=target_sum, inplace=False)['X']
        z_scores = sc.pp.log1p(z_scores)
        if isinstance(target_variance, numbers.Number):
            z_scores = sc.pp.scale(z_scores, max_value=max_value)
            z_scores *= np.sqrt(target_variance)
        else:
            z_scores = sc.pp.scale(z_scores)
            z_scores[z_scores < max_value] = max_value
            z_scores *= np.sqrt(target_variance)

        for res in results_list:
            res.calculate_gene_coefs(z_scores, adata.var_names)

    @staticmethod
    def _run_nmf(x, nmf_kwargs, name, tens: 'torch.Tensor' = None, verbose: int = 0) -> NMFResult:
        if tens is not None:
            W, H, _ = cnmf.nmf_torch(x, nmf_kwargs, tens, verbose=(verbose > 0))
        else:
            W, H, _ = sknmf.non_negative_factorization(x, **nmf_kwargs, verbose=verbose)

        # X ~ W @ H, transpose for cells to be columns
        loss_per_cell = pfnmf.calc_beta_divergence(
            x.T, H.T, np.zeros((H.shape[1], 0)),
            W.T, np.zeros((0, W.shape[0])), per_column=True)

        return NMFResult(
            name=name,
            loss_per_cell=loss_per_cell,
            rank=W.shape[1],
            W=W,
            H=H)

    @staticmethod
    def _get_ordered_adjacency_matrix(correlation_matrix, prog_names, ranks, threshold=0.2, verbose: bool = False):
        """
        Given a correlation matrix to base the adjacency matrix on, returns the
        adjacency matrix after filtering out edges with correlation below the threshold
        and keeping only edges between consecutive layers.
        """
        # adjacency matrix creation
        adjacency = pd.DataFrame(np.round((correlation_matrix), 2),
                              index=prog_names, columns=prog_names)

        # order
        linkage = hierarchy.linkage(
            adjacency, method='average', metric='euclidean')
        prog_order = hierarchy.leaves_list(
            hierarchy.optimal_leaf_ordering(linkage, adjacency))

        # keeping only edges between consecutive layers
        for i in range(len(ranks) - 2):
            adjacency.values[:np.sum(ranks[:i + 1]), np.sum(ranks[:i + 2]):] = 0
            adjacency.values[np.sum(ranks[:i + 2]):, :np.sum(ranks[:i + 1])] = 0

        np.fill_diagonal(adjacency.values, 0)
        adjacency.values[adjacency.values <= threshold] = 0

        if verbose:
            print(f'Number of edges={np.count_nonzero(adjacency)}')

        # ordering the nodes for display
        adjacency = adjacency.iloc[prog_order, prog_order]

        return adjacency

    @staticmethod
    def _plot_layered_correlation_flow_chart(layer_keys, adjacency_df: pd.DataFrame,
                                            prog_names_dict, title: str,
                                            plt_figure_kwargs: Dict = None,
                                            fig_title_kwargs: Dict = None):
        """
        Plotting the flow chart of the correlation matrix between layers.
        """
        # setting figure arguments
        figure_kwargs = {'figsize': (14.4, 16.2), 'dpi': 100}
        if plt_figure_kwargs is not None: figure_kwargs.update(plt_figure_kwargs)

        title_kwargs = {'fontsize': 25, 'y': 0.95}
        if fig_title_kwargs is not None: title_kwargs.update(fig_title_kwargs)

        # mapping adata short name to layer number
        name_map = dict(zip(layer_keys, range(len(layer_keys))))

        # create the graph object
        G = nx.from_numpy_array(adjacency_df.values, create_using=nx.Graph)
        nx.relabel_nodes(G, lambda i: adjacency_df.index[i], copy=False)
        nx.set_node_attributes(
            G, {node: name_map[node.split('.')[0]] for node in G.nodes}, name='layer')

        # prepare graph for display
        layout = nx.multipartite_layout(G, subset_key='layer')

        edges, weights = zip(*nx.get_edge_attributes(G, 'weight').items())
        edge_width = 15 * np.power(weights, 2)  # visual edge emphesis

        if len(layer_keys) > 2:
            for layer in {data['layer'] for key, data in G.nodes.data()}:
                nodes = [node for node in G.nodes if name_map[node.split('.')[0]] == layer]

                angles = np.linspace(-np.pi / 4, np.pi / 4, len(nodes))

                for i, node in enumerate(nodes):
                    layout[node] = [layer + 2 * np.cos(angles[i]), np.sin(angles[i])]

        fig, ax = plt.subplots(1, 1, **figure_kwargs)
        nx.draw(G, layout, node_size=3000, with_labels=False, edge_color=weights,
                edge_vmin=0, edge_vmax=1., width=edge_width, ax=ax)

        cmp = plt.matplotlib.cm.ScalarMappable(plt.matplotlib.colors.Normalize(vmin=0, vmax=1))
        plt.colorbar(cmp, orientation='horizontal', cax=fig.add_subplot(18, 5, 86))

        # change color of layers
        for key in layer_keys:
            nx.draw_networkx_nodes(
                G, layout, node_size=2800, nodelist=prog_names_dict[key],
                # node_color=coloring_scheme[key],
                ax=ax)
        nx.draw_networkx_labels(G, layout, font_size=11, ax=ax)

        plt.suptitle(title, **title_kwargs)

        plt.tight_layout()
        return fig

